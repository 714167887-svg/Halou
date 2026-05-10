"""
KB 自动排版（基于 PB 提取规则）

目标：
1) 默认读取 W/PB/板材尺寸.xlsx（颜色/厚度/高度）
2) 固定板宽 1219
3) 从 CAD（11/*.dwg|*.dxf）按订单号匹配排版
4) 同一订单号尽量放在同一块板；若同订单存在不同厚度，会分到不同材质文件
5) 输出文件命名：{颜色}{厚度}.dxf（如：枪灰拉丝0.71.dxf）

用法：
  python kb_layout.py
"""

from __future__ import annotations

import os
import re
import sys
import math
import shutil
from dataclasses import dataclass, field
from typing import Dict, List, Tuple, Optional

import ezdxf
from openpyxl import load_workbook

import pb_extract

BOARD_WIDTH = 1219.0
PIECE_GAP = 8.0
ROW_GAP = 8.0
PLATE_GAP = 300.0


@dataclass
class Piece:
	order_no: str
	order_group: str
	color: str
	thickness: float
	block_name: str
	width: float
	height: float
	source_file: str


@dataclass
class PlateState:
	x: float = 0.0
	y: float = 0.0
	row_h: float = 0.0
	placements: List[Tuple[Piece, float, float]] = field(default_factory=list)


def _norm_text(s: Optional[str]) -> str:
	if s is None:
		return ""
	return str(s).strip()


def _norm_color(s: str) -> str:
	return re.sub(r"\s+", "", _norm_text(s))


def _thickness_str(v: float) -> str:
	s = f"{v:.2f}".rstrip("0").rstrip(".")
	return s if s else "0"


def _extract_order_group(order_no: str) -> str:
	order_no = _norm_text(order_no)
	return re.sub(r"-\d+$", "", order_no)


def load_board_catalog(xlsx_path: str) -> List[Tuple[str, float, float]]:
	if not os.path.exists(xlsx_path):
		raise FileNotFoundError(f"未找到板材数据文件: {xlsx_path}")

	wb = load_workbook(xlsx_path, data_only=True)
	ws = wb[wb.sheetnames[0]]
	headers = [_norm_text(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
	idx = {h: i + 1 for i, h in enumerate(headers) if h}

	for req in ("颜色", "厚度", "高度"):
		if req not in idx:
			raise ValueError(f"板材数据缺少列: {req}，当前表头: {headers}")

	rows: List[Tuple[str, float, float]] = []
	for r in range(2, ws.max_row + 1):
		color = _norm_color(ws.cell(r, idx["颜色"]).value)
		t = ws.cell(r, idx["厚度"]).value
		h = ws.cell(r, idx["高度"]).value
		if not color or t is None or h is None:
			continue
		try:
			rows.append((color, float(t), float(h)))
		except Exception:
			continue

	if not rows:
		raise ValueError("板材数据为空：未读取到有效的 颜色/厚度/高度 行")

	return rows


def find_board_height(catalog: List[Tuple[str, float, float]], color: str, thickness: float) -> float:
	c = _norm_color(color)
	t = float(thickness)

	cand = [h for cc, tt, h in catalog if cc == c and abs(tt - t) <= 0.03]
	if cand:
		return max(cand)

	cand = [h for cc, tt, h in catalog if (cc in c or c in cc) and abs(tt - t) <= 0.03]
	if cand:
		return max(cand)

	cand = [h for _, tt, h in catalog if abs(tt - t) <= 0.03]
	if cand:
		return max(cand)

	return max(h for _, _, h in catalog)


def detect_color_from_text(text: str, known_colors: List[str]) -> str:
	t = _norm_text(text)
	if not t:
		return ""
	for c in known_colors:
		if c and c in t:
			return c
	return ""


def _point_in_rect(x: float, y: float, x1: float, x2: float, y1: float, y2: float, tol: float = 2.0) -> bool:
	return (x1 - tol) <= x <= (x2 + tol) and (y1 - tol) <= y <= (y2 + tol)


def _collect_zone_texts(msp, zone: dict, tol: float = 2.0) -> List[str]:
	x1 = zone["box"]["min_x"]
	x2 = zone["box"]["max_x"]
	y1 = zone["mag_y"]
	y2 = zone["box"]["max_y"]

	out: List[str] = []
	for e in msp:
		try:
			t = e.dxftype()
			if t == "TEXT":
				ix, iy = e.dxf.insert.x, e.dxf.insert.y
				if _point_in_rect(ix, iy, x1, x2, y1, y2, tol):
					s = _norm_text(e.dxf.text)
					if s:
						out.append(s)
			elif t == "MTEXT":
				ix, iy = e.dxf.insert.x, e.dxf.insert.y
				if _point_in_rect(ix, iy, x1, x2, y1, y2, tol):
					s = _norm_text(e.plain_text())
					if s:
						out.append(s)
		except Exception:
			continue
	return out


def parse_zone_meta(texts: List[str], file_name: str, known_colors: List[str]) -> Tuple[str, str, float]:
	joined = "\n".join(texts)

	m_order = re.search(r"ERP单号\s*[；;:：]?\s*([A-Za-z0-9\-]+)", joined, flags=re.IGNORECASE)
	if m_order:
		order_no = m_order.group(1).strip()
	else:
		stem = os.path.splitext(file_name)[0]
		m_file = re.search(r"-([0-9A-Za-z]{3,})-", stem)
		order_no = m_file.group(1) if m_file else stem

	m_color = re.search(r"颜色\s*[；;:：]?\s*([^\s；;\n\r]+)", joined)
	color = m_color.group(1).strip() if m_color else "未识别颜色"
	if color == "未识别颜色":
		guessed = detect_color_from_text(joined, known_colors)
		if guessed:
			color = guessed

	thks: List[float] = []
	for m in re.finditer(r"(\d+(?:\.\d+)?)\s*/\s*\d+(?:\.\d+)?", joined):
		v = float(m.group(1))
		if 0.3 <= v <= 3.0:
			thks.append(v)

	if not thks:
		for m in re.finditer(r"厚度\s*[；;:：]?\s*(\d+(?:\.\d+)?)", joined):
			v = float(m.group(1))
			if 0.3 <= v <= 3.0:
				thks.append(v)

	thickness = thks[0] if thks else 0.76
	return order_no, color, thickness


def collect_doc_meta_anchors(msp, known_colors: List[str]) -> List[Tuple[float, float, str, str]]:
	anchors: List[Tuple[float, float, str, str]] = []
	for e in msp:
		try:
			t = e.dxftype()
			if t == "TEXT":
				txt = _norm_text(e.dxf.text)
				ix, iy = e.dxf.insert.x, e.dxf.insert.y
			elif t == "MTEXT":
				txt = _norm_text(e.plain_text())
				ix, iy = e.dxf.insert.x, e.dxf.insert.y
			else:
				continue

			if not txt:
				continue

			m_order = re.search(r"ERP单号\s*[；;:：]?\s*([A-Za-z0-9\-]+)", txt, flags=re.IGNORECASE)
			m_color = re.search(r"颜色\s*[；;:：]?\s*([^\s；;\n\r]+)", txt)
			guessed_color = detect_color_from_text(txt, known_colors)
			if not (m_order or m_color or guessed_color):
				continue

			order_no = m_order.group(1).strip() if m_order else ""
			color = m_color.group(1).strip() if m_color else guessed_color
			anchors.append((ix, iy, order_no, color))
		except Exception:
			continue

	return anchors


def nearest_anchor_meta(zone: dict, anchors: List[Tuple[float, float, str, str]]) -> Tuple[str, str]:
	if not anchors:
		return "", ""

	cx = (zone["box"]["min_x"] + zone["box"]["max_x"]) / 2.0
	cy = (zone["box"]["min_y"] + zone["box"]["max_y"]) / 2.0

	best = None
	best_d2 = None
	for ax, ay, order_no, color in anchors:
		d2 = (ax - cx) ** 2 + (ay - cy) ** 2
		if best_d2 is None or d2 < best_d2:
			best_d2 = d2
			best = (order_no, color)

	return best if best else ("", "")


def _entity_bbox(e) -> Optional[Tuple[float, float, float, float]]:
	try:
		t = e.dxftype()
		if t == "LINE":
			x1, y1 = e.dxf.start.x, e.dxf.start.y
			x2, y2 = e.dxf.end.x, e.dxf.end.y
			return min(x1, x2), min(y1, y2), max(x1, x2), max(y1, y2)
		if t == "LWPOLYLINE":
			pts = list(e.get_points(format="xy"))
			if not pts:
				return None
			xs = [p[0] for p in pts]
			ys = [p[1] for p in pts]
			return min(xs), min(ys), max(xs), max(ys)
		if t in ("CIRCLE", "ARC"):
			cx, cy, r = e.dxf.center.x, e.dxf.center.y, e.dxf.radius
			return cx - r, cy - r, cx + r, cy + r
		if t == "INSERT":
			ix, iy = e.dxf.insert.x, e.dxf.insert.y
			return ix, iy, ix, iy
	except Exception:
		return None
	return None


def _collect_piece_bounds(entities: List) -> Tuple[float, float, float, float]:
	mins = []
	for e in entities:
		b = _entity_bbox(e)
		if b is not None:
			mins.append(b)

	if not mins:
		return 0.0, 0.0, 10.0, 10.0

	min_x = min(v[0] for v in mins)
	min_y = min(v[1] for v in mins)
	max_x = max(v[2] for v in mins)
	max_y = max(v[3] for v in mins)
	return min_x, min_y, max_x, max_y


def _simulate_place(state: PlateState, pw: float, ph: float, board_h: float) -> Optional[Tuple[float, float, float, float, float]]:
	x, y, row_h = state.x, state.y, state.row_h

	if pw > BOARD_WIDTH or ph > board_h:
		return None

	if x > 0 and x + pw > BOARD_WIDTH:
		x = 0.0
		y += row_h + ROW_GAP
		row_h = 0.0

	if y + ph > board_h:
		return None

	px, py = x, y
	x = x + pw + PIECE_GAP
	row_h = max(row_h, ph)
	return px, py, x, y, row_h


def _try_place_order_together(state: PlateState, pieces: List[Piece], board_h: float) -> Optional[List[Tuple[Piece, float, float]]]:
	x, y, row_h = state.x, state.y, state.row_h
	placements: List[Tuple[Piece, float, float]] = []

	for p in pieces:
		if p.width > BOARD_WIDTH or p.height > board_h:
			return None

		tx, ty, tr = x, y, row_h
		if tx > 0 and tx + p.width > BOARD_WIDTH:
			tx = 0.0
			ty += tr + ROW_GAP
			tr = 0.0

		if ty + p.height > board_h:
			return None

		placements.append((p, tx, ty))
		x = tx + p.width + PIECE_GAP
		y = ty
		row_h = max(tr, p.height)

	return placements


def _commit_order(state: PlateState, placements: List[Tuple[Piece, float, float]]):
	for p, px, py in placements:
		state.placements.append((p, px, py))

	if not placements:
		return

	last_p, last_x, last_y = placements[-1]
	state.x = last_x + last_p.width + PIECE_GAP
	state.y = last_y
	state.row_h = max((p.height for p, _, py in state.placements if abs(py - last_y) < 1e-6), default=0.0)


def _add_plate_outline(msp, ox: float, board_h: float):
	pts = [(ox, 0), (ox + BOARD_WIDTH, 0), (ox + BOARD_WIDTH, board_h), (ox, board_h)]
	pl = msp.add_lwpolyline(pts, close=True)
	pl.dxf.color = 3


def build_layout_for_group(doc: ezdxf.EzDxf, pieces: List[Piece], color: str, thickness: float, board_h: float):
	msp = doc.modelspace()

	by_order: Dict[str, List[Piece]] = {}
	for p in pieces:
		by_order.setdefault(p.order_group, []).append(p)

	for k in by_order:
		by_order[k].sort(key=lambda it: it.width * it.height, reverse=True)

	order_keys = sorted(by_order.keys())
	plates: List[PlateState] = [PlateState()]
	current_plate = plates[0]

	for ok in order_keys:
		order_pieces = by_order[ok]
		sim = _try_place_order_together(current_plate, order_pieces, board_h)
		if sim is not None:
			_commit_order(current_plate, sim)
			continue

		current_plate = PlateState()
		plates.append(current_plate)
		sim = _try_place_order_together(current_plate, order_pieces, board_h)
		if sim is not None:
			_commit_order(current_plate, sim)
			continue

		for p in order_pieces:
			placed = False
			trial = _simulate_place(current_plate, p.width, p.height, board_h)
			if trial is not None:
				px, py, nx, ny, nrh = trial
				current_plate.placements.append((p, px, py))
				current_plate.x, current_plate.y, current_plate.row_h = nx, ny, nrh
				placed = True

			if not placed:
				current_plate = PlateState()
				plates.append(current_plate)
				trial = _simulate_place(current_plate, p.width, p.height, board_h)
				if trial is None:
					print(
						f"[WARN] 零件过大无法排入板材: order={p.order_no}, {p.width:.1f}x{p.height:.1f}, 板={BOARD_WIDTH}x{board_h}"
					)
					continue
				px, py, nx, ny, nrh = trial
				current_plate.placements.append((p, px, py))
				current_plate.x, current_plate.y, current_plate.row_h = nx, ny, nrh

	for i, plate in enumerate(plates, start=1):
		ox = (i - 1) * (BOARD_WIDTH + PLATE_GAP)
		_add_plate_outline(msp, ox, board_h)

		for p, x, y in plate.placements:
			msp.add_blockref(p.block_name, (ox + x, y))
			tx = ox + x
			ty = y + p.height + 5
			msp.add_text(p.order_no, dxfattribs={"height": 8}).set_placement((tx, ty))

		title = f"板{i}  {color}{_thickness_str(thickness)}  {BOARD_WIDTH}x{int(board_h)}"
		msp.add_text(title, dxfattribs={"height": 12, "color": 5}).set_placement((ox, board_h + 20))


def sanitize_filename(name: str) -> str:
	return re.sub(r"[\\/:*?\"<>|]", "_", name)


def main():
	base_dir = os.path.dirname(os.path.abspath(__file__))
	input_dir = os.path.join(base_dir, "11")
	board_xlsx = os.path.join(base_dir, "板材尺寸.xlsx")
	out_dir = os.path.join(base_dir, "KB_output")
	os.makedirs(out_dir, exist_ok=True)

	for name in os.listdir(out_dir):
		if name.lower().endswith(".dxf"):
			try:
				os.remove(os.path.join(out_dir, name))
			except OSError:
				pass

	catalog = load_board_catalog(board_xlsx)
	known_colors = sorted({_norm_color(c) for c, _, _ in catalog if _norm_color(c)}, key=len, reverse=True)
	print(f"板材数据加载完成：{len(catalog)} 条（来源：{board_xlsx}）")

	cad_files = sorted(f for f in os.listdir(input_dir) if f.lower().endswith((".dwg", ".dxf")))
	if not cad_files:
		print(f"未找到输入 CAD：{input_dir}")
		sys.exit(1)

	groups = {}
	total_extracted_pieces = 0

	for cad_name in cad_files:
		cad_path = os.path.join(input_dir, cad_name)
		print(f"\n{'=' * 70}\n处理: {cad_name}")

		tmp_dir = None
		try:
			if cad_name.lower().endswith(".dwg"):
				dxf_path, tmp_dir = pb_extract.convert_dwg_to_dxf(cad_path)
			else:
				dxf_path = cad_path

			src_doc = ezdxf.readfile(dxf_path)
			src_msp = src_doc.modelspace()
			zones = pb_extract.find_cyan_boxes_with_magenta(src_msp)
			anchors = collect_doc_meta_anchors(src_msp, known_colors)
			print(f"  zones={len(zones)}")

			if not zones:
				continue

			importers = {}

			for zi, zone in enumerate(zones, start=1):
				texts = _collect_zone_texts(src_msp, zone)
				order_no, color, thickness = parse_zone_meta(texts, cad_name, known_colors)

				anchor_order, anchor_color = nearest_anchor_meta(zone, anchors)
				if (not order_no) or (order_no in ("", os.path.splitext(cad_name)[0])):
					if anchor_order:
						order_no = anchor_order
				if color == "未识别颜色" and anchor_color:
					color = anchor_color

				order_group = _extract_order_group(order_no)

				key = (_norm_color(color), _thickness_str(thickness))
				if key not in groups:
					out_doc = ezdxf.new(dxfversion="R2010")
					groups[key] = {
						"doc": out_doc,
						"pieces": [],
						"color": _norm_color(color),
						"thickness": float(thickness),
					}

				group = groups[key]
				out_doc = group["doc"]

				entities = [e for e in src_msp if pb_extract.entity_in_zone(e, zone)]
				if not entities:
					print(f"  zone#{zi} 无可提取实体，跳过")
					continue

				min_x, min_y, max_x, max_y = _collect_piece_bounds(entities)
				w = max_x - min_x
				h = max_y - min_y
				is_landscape = w > h

				from ezdxf.math import Matrix44

				if is_landscape:
					m_to_origin = Matrix44.translate(-min_x, -min_y, 0)
					m_rotate = Matrix44.z_rotate(math.pi / 2)
					m_shift = Matrix44.translate(h, 0, 0)
					tfm = m_to_origin @ m_rotate @ m_shift
					pw, ph = h, w
				else:
					tfm = Matrix44.translate(-min_x, -min_y, 0)
					pw, ph = w, h

				safe_src = re.sub(r"[^A-Za-z0-9_\-]", "_", os.path.splitext(cad_name)[0])
				block_name = f"KB_{safe_src}_{zi}_{len(group['pieces']) + 1}"
				while block_name in out_doc.blocks:
					block_name += "_x"
				blk = out_doc.blocks.new(name=block_name)

				if key not in importers:
					from ezdxf.addons import Importer

					importer = Importer(src_doc, out_doc)
					importer.import_tables()
					importers[key] = importer
				importer = importers[key]

				for src in entities:
					copied = None
					try:
						copied = src.copy()
						src_msp.add_entity(copied)
						copied.transform(tfm)
						importer.import_entity(copied, blk)
					except Exception:
						pass
					finally:
						if copied is not None:
							try:
								src_msp.delete_entity(copied)
							except Exception:
								pass

				group["pieces"].append(
					Piece(
						order_no=order_no,
						order_group=order_group,
						color=_norm_color(color),
						thickness=float(thickness),
						block_name=block_name,
						width=max(pw, 1.0),
						height=max(ph, 1.0),
						source_file=cad_name,
					)
				)
				total_extracted_pieces += 1
				print(f"  zone#{zi}: order={order_no}, color={color}, thk={thickness:.2f}, size={pw:.1f}x{ph:.1f}")

			for imp in importers.values():
				imp.finalize()

		finally:
			if tmp_dir and os.path.exists(tmp_dir):
				shutil.rmtree(tmp_dir, ignore_errors=True)

	if not groups:
		print("未生成任何有效分组，结束。")
		sys.exit(1)

	print(f"\n开始输出分组文件：{len(groups)} 组")
	for key, info in groups.items():
		color = info["color"]
		thickness = float(info["thickness"])
		pieces: List[Piece] = info["pieces"]
		doc = info["doc"]

		board_h = find_board_height(catalog, color, thickness)
		build_layout_for_group(doc, pieces, color, thickness, board_h)

		out_name = sanitize_filename(f"{color}{_thickness_str(thickness)}.dxf")
		out_path = os.path.join(out_dir, out_name)
		doc.saveas(out_path)

		print(f"  已输出: {out_name} | 零件={len(pieces)} | 板材={BOARD_WIDTH}x{board_h:.0f}")

	print(f"\n总抓取零件数: {total_extracted_pieces}")
	print(f"完成。输出目录: {out_dir}")


if __name__ == "__main__":
	main()
